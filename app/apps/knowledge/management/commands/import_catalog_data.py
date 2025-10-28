import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise CommandError("The command requires openpyxl. Please install it in the environment.") from exc

from knowledge.models import (
    EraReference,
    GazetteerStructureRecord,
    LibraryCatalog,
    PersonReference,
    PlaceReference,
)


class Command(BaseCommand):
    help = "Import catalog and reference data from Excel/CSV files into the knowledge module."

    def add_arguments(self, parser):
        parser.add_argument("--catalog", type=str, help="Path to 温州古旧方志目录.xlsx (or CSV export).")
        parser.add_argument("--structure", type=str, help="Path to 乾隆温州府志卷首结构 Excel.")
        parser.add_argument("--places", type=str, help="Path to 明清地名志数据表 Excel.")
        parser.add_argument("--eras", type=str, help="Path to 年号对照数据 (CSV/Excel).")
        parser.add_argument("--persons", type=str, help="Path to 人物基础数据 (CSV/Excel).")
        parser.add_argument(
            "--format",
            dest="file_format",
            choices=("excel", "csv"),
            default="excel",
            help="Input file format (default: excel).",
        )

    def handle(self, *args, **options):
        file_format = options["file_format"]
        if file_format not in {"excel", "csv"}:
            raise CommandError("Unsupported file format. Use 'excel' or 'csv'.")

        total = 0
        if options.get("catalog"):
            total += self.import_catalog(Path(options["catalog"]), file_format)

        if options.get("structure"):
            total += self.import_structure(Path(options["structure"]), file_format)

        if options.get("places"):
            total += self.import_places(Path(options["places"]), file_format)

        if options.get("eras"):
            total += self.import_eras(Path(options["eras"]), file_format)

        if options.get("persons"):
            total += self.import_persons(Path(options["persons"]), file_format)

        if not total:
            raise CommandError("No data imported. Provide at least one file via --catalog/--structure/--places.")

        self.stdout.write(self.style.SUCCESS(f"Import completed. {total} records processed."))

    # region helpers
    def import_catalog(self, path: Path, file_format: str) -> int:
        records = list(self.iter_rows(path, file_format))
        if not records:
            self.stdout.write("No rows found in catalog file.")
            return 0

        count = 0
        for row in records:
            payload = {
                "category": row.get("分类", "") or "",
                "title": row.get("书名", "") or "",
                "author": row.get("著者", "") or "",
                "edition": row.get("版本", "") or "",
                "volume_count": stringify(row.get("册数")),
                "collection_location": row.get("馆藏", "") or "",
                "call_number": row.get("索书号", "") or "",
                "page_count": stringify(row.get("页数")),
                "source_filename": path.name,
                "extra": {k: v for k, v in row.items() if k not in {"分类", "书名", "著者", "版本", "册数", "馆藏", "索书号", "页数"}},
            }
            if not payload["title"]:
                continue
            lookup = {"title": payload["title"], "call_number": payload["call_number"]}
            if not payload["call_number"]:
                lookup = {"title": payload["title"], "edition": payload["edition"]}

            LibraryCatalog.objects.update_or_create(defaults=payload, **lookup)
            count += 1

        self.stdout.write(self.style.NOTICE(f"Catalog: processed {count} rows from {path}"))
        return count

    def import_structure(self, path: Path, file_format: str) -> int:
        records = list(self.iter_rows(path, file_format))
        if not records:
            self.stdout.write("No rows found in structure file.")
            return 0

        count = 0
        for row in records:
            record_id = row.get("ID") or row.get("唯一标识符号")
            if not record_id:
                continue

            payload = {
                "dataset": row.get("史籍语料", "") or "",
                "record_id": record_id,
                "unique_identifier": row.get("唯一标识符号", "") or "",
                "title_level": row.get("标题级别", "") or "",
                "new_title_level": row.get("新增标题级别", "") or "",
                "extracted_title": row.get("提取标题", "") or "",
                "subject_terms": serialize_list(row.get("主题词")),
                "main_responsible": row.get("主要责任者", "") or "",
                "abstract": row.get("摘要", "") or "",
                "funding": row.get("资助", "") or "",
                "related_resources": row.get("相关资源", "") or "",
                "other_language_title": row.get("其他语种题名", "") or "",
                "other_language_subject_terms": serialize_list(row.get("其他语种主题词")),
                "other_language_abstract": row.get("其他语种摘要", "") or "",
                "other_language_funding": row.get("其他语种资助", "") or "",
                "language": row.get("语种", "") or "",
                "classification_ccl": row.get("中图分类法", "") or "",
                "academic_classification": row.get("学科分类法", "") or "",
                "industry_classification": row.get("行业分类法", "") or "",
                "era_classification": row.get("时代分类法", "") or "",
                "region_classification": row.get("地区分类法", "") or "",
                "column": row.get("专栏", "") or "",
                "source_filename": path.name,
                "extra": {k: v for k, v in row.items() if k not in {
                    "史籍语料", "ID", "唯一标识符号", "标题级别", "新增标题级别", "提取标题", "主题词",
                    "主要责任者", "摘要", "资助", "相关资源", "其他语种题名", "其他语种主题词", "其他语种摘要",
                    "其他语种资助", "语种", "中图分类法", "学科分类法", "行业分类法", "时代分类法",
                    "地区分类法", "专栏"
                }},
            }

            GazetteerStructureRecord.objects.update_or_create(record_id=record_id, defaults=payload)
            count += 1

        self.stdout.write(self.style.NOTICE(f"Structure: processed {count} rows from {path}"))
        return count

    def import_places(self, path: Path, file_format: str) -> int:
        records = list(self.iter_rows(path, file_format))
        if not records:
            self.stdout.write("No rows found in places file.")
            return 0

        count = 0
        for row in records:
            standard_name = row.get("標準地名") or row.get("标准地名") or row.get("标准名称")
            if not standard_name:
                continue

            payload = {
                "dynasty": row.get("朝代", "") or "",
                "standard_name": standard_name,
                "admin_level": row.get("行政層級", "") or row.get("行政层级", "") or "",
                "level_description": row.get("層級說明", "") or row.get("层级说明", "") or "",
                "era_years_chinese": row.get("沿革年代(中曆)", "") or "",
                "era_years_western": row.get("沿革年代(西曆)", "") or "",
                "alternate_names": serialize_list(row.get("異名")),
                "event_codes": row.get("事件代碼", "") or "",
                "evolution_notes": row.get("沿革說明", "") or "",
                "affiliation": row.get("隸屬", "") or "",
                "affiliation_level_description": row.get("層級說明(隸屬)", "") or "",
                "jurisdiction": serialize_list(row.get("轄區")),
                "center_longitude": parse_float(row.get("中心經度")),
                "center_latitude": parse_float(row.get("中心緯度")),
                "east_longitude": parse_float(row.get("東端經度")),
                "west_longitude": parse_float(row.get("西端經度")),
                "south_latitude": parse_float(row.get("南端緯度")),
                "north_latitude": parse_float(row.get("北端緯度")),
                "east_neighbor": row.get("東鄰地名", "") or "",
                "west_neighbor": row.get("西鄰地名", "") or "",
                "south_neighbor": row.get("南鄰地名", "") or "",
                "north_neighbor": row.get("北鄰地名", "") or "",
                "neighbors_era": row.get("四鄰年代", "") or "",
                "references": row.get("參考文獻", "") or "",
                "notes": row.get("備註", "") or "",
                "source_filename": path.name,
                "extra": {k: v for k, v in row.items() if k not in {
                    "朝代", "標準地名", "行政層級", "層級說明", "沿革年代(中曆)", "沿革年代(西曆)", "異名",
                    "事件代碼", "沿革說明", "隸屬", "層級說明(隸屬)", "轄區", "中心經度", "中心緯度",
                    "東端經度", "西端經度", "南端緯度", "北端緯度", "東鄰地名", "西鄰地名", "南鄰地名",
                    "北鄰地名", "四鄰年代", "參考文獻", "備註"
                }},
            }

            lookup = {"standard_name": payload["standard_name"], "dynasty": payload["dynasty"]}
            PlaceReference.objects.update_or_create(defaults=payload, **lookup)
            count += 1

        self.stdout.write(self.style.NOTICE(f"Places: processed {count} rows from {path}"))
        return count

    def import_eras(self, path: Path, file_format: str) -> int:
        records = list(self.iter_rows(path, file_format))
        if not records:
            self.stdout.write("No rows found in eras file.")
            return 0

        count = 0
        for row in records:
            era_id = stringify(row.get("era_id") or row.get("Era ID"))
            era_name = stringify(row.get("era_name") or row.get("Era Name"))
            if not era_id or not era_name:
                continue

            payload = {
                "era_id": era_id,
                "era_name": era_name,
                "dynasty": stringify(row.get("dynasty") or row.get("Dynasty")),
                "emperor": stringify(row.get("emperor") or row.get("Emperor")),
                "start_year_ce": parse_int(row.get("start_year_ce") or row.get("Start Year CE")),
                "end_year_ce": parse_int(row.get("end_year_ce") or row.get("End Year CE")),
                "start_year_cn": stringify(row.get("start_year_cn") or row.get("Start Year CN")),
                "end_year_cn": stringify(row.get("end_year_cn") or row.get("End Year CN")),
                "applicable_regions": stringify(row.get("applicable_regions") or row.get("Applicable Regions")),
                "source_refs": serialize_list_to_list(row.get("source_refs") or row.get("Source Refs")),
                "notes": stringify(row.get("notes") or row.get("Notes")),
            }
            if payload["start_year_ce"] is None or payload["end_year_ce"] is None:
                self.stdout.write(self.style.WARNING(f"Skipping era {era_id}: missing start/end year."))
                continue

            EraReference.objects.update_or_create(era_id=era_id, defaults=payload)
            count += 1

        self.stdout.write(self.style.NOTICE(f"Eras: processed {count} rows from {path}"))
        return count

    def import_persons(self, path: Path, file_format: str) -> int:
        records = list(self.iter_rows(path, file_format))
        if not records:
            self.stdout.write("No rows found in persons file.")
            return 0

        count = 0
        for row in records:
            person_id = stringify(row.get("person_id") or row.get("Person ID"))
            name = stringify(row.get("name") or row.get("Name"))
            if not person_id or not name:
                continue

            origin_place = None
            place_lookup = stringify(row.get("origin_place_id") or row.get("Origin Place ID"))
            place_name = stringify(row.get("origin_place_name") or row.get("Origin Place"))
            if place_lookup:
                place_pk = parse_int(place_lookup)
                if place_pk is not None:
                    origin_place = PlaceReference.objects.filter(pk=place_pk).first()
            if origin_place is None and place_name:
                origin_place = PlaceReference.objects.filter(standard_name=place_name).first()

            payload = {
                "person_id": person_id,
                "name": name,
                "courtesy_name": stringify(row.get("courtesy_name") or row.get("Courtesy Name")),
                "aliases": serialize_list_to_list(row.get("aliases") or row.get("Aliases")),
                "gender": stringify(row.get("gender") or row.get("Gender")),
                "dynasty": stringify(row.get("dynasty") or row.get("Dynasty")),
                "birth_year": parse_int(row.get("birth_year") or row.get("Birth Year")),
                "death_year": parse_int(row.get("death_year") or row.get("Death Year")),
                "origin_place": origin_place,
                "positions": serialize_list_to_list(row.get("positions") or row.get("Positions")),
                "works": serialize_list_to_list(row.get("works") or row.get("Works")),
                "related_events": serialize_list_to_list(row.get("related_events") or row.get("Related Events")),
                "biography_summary": stringify(row.get("biography_summary") or row.get("Biography Summary")),
                "source_refs": serialize_list_to_list(row.get("source_refs") or row.get("Source Refs")),
                "notes": stringify(row.get("notes") or row.get("Notes")),
            }

            PersonReference.objects.update_or_create(person_id=person_id, defaults=payload)
            count += 1

        self.stdout.write(self.style.NOTICE(f"Persons: processed {count} rows from {path}"))
        return count

    def iter_rows(self, path: Path, file_format: str):
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        if file_format == "csv":
            with path.open(newline="", encoding="utf-8-sig") as handle:
                reader = csv.DictReader(handle)
                yield from reader
        else:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = None
            for row in rows:
                if headers is None:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                values = {headers[idx]: (row[idx] if idx < len(row) else None) for idx in range(len(headers))}
                if not any(values.values()):
                    continue
                yield values
            workbook.close()


def stringify(value):
    if value is None:
        return ""
    return str(value).strip()


def serialize_list(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(stringify(v) for v in value if stringify(v))
    text = stringify(value)
    # Some cells may use newline separators
    normalized = text.replace("\n", ",").replace(";", ",")
    parts = [part.strip() for part in normalized.split(",") if part and part.strip()]
    return ", ".join(parts)


def serialize_list_to_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [stringify(v) for v in value if stringify(v)]
    text = stringify(value)
    normalized = text.replace("\n", ",").replace(";", ",")
    parts = [part.strip() for part in normalized.split(",") if part and part.strip()]
    return parts


def parse_float(value):
    try:
        if value in ("", None):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_int(value):
    try:
        if value in ("", None):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None
